"""
India Power Analytics — Supabase Uploader
==========================================
Parses and uploads data from all confirmed India power data sources.

DATA SOURCES (priority order):
  P1 Daily:  CEA OPM Daily Generation Report
             cea.nic.in/opm_grid_operation/daily-generation-report/
             Files: All_India_Generation_YYYY-MM-DD.csv
                    Demand_Met_Data_YYYY-MM-DD.csv

  P1 Daily:  CEA RE Daily Report (RPMD — Renewable Project Monitoring Div)
             cea.nic.in/renewable-generation-report/
             File:  cea_re_YYYY-MM-DD.pdf

  P2 Daily:  NPP dgr2 — Station/Unit Wise (most comprehensive, 382 KB)
             npp.gov.in/public-reports/cea/daily/dgr/DD-MM-YYYY/dgr2-YYYY-MM-DD.xls

  P2 Daily:  NPP dgr3 — All India Fuel-wise Summary (smallest, 9 KB)
             npp.gov.in/public-reports/cea/daily/dgr/DD-MM-YYYY/dgr3-YYYY-MM-DD.xls

USAGE:
  # Upload CEA OPM CSVs (standard daily workflow)
  python upload_to_supabase.py --date 2026-06-20 --gen All_India_Generation_2026-06-20.csv --dem Demand_Met_Data_2026-06-20.csv

  # Upload CEA RE PDF (state-level wind/solar breakdown)
  python upload_to_supabase.py --date 2026-06-20 --re cea_re_2026-06-20.pdf

  # Upload NPP dgr2 XLS (station-level — most comprehensive)
  python upload_to_supabase.py --date 2026-06-20 --dgr2 dgr2-2026-06-20.xls

  # Upload everything at once
  python upload_to_supabase.py --date 2026-06-20 --auto

REQUIREMENTS:
  pip install supabase pandas pdfplumber openpyxl
"""

import sys
import re
import logging
import argparse
from pathlib import Path
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

try:
    import pandas as pd
    PD_OK = True
except ImportError:
    PD_OK = False
    print("[ERROR] pandas not installed. Run: pip install pandas")

try:
    from supabase import create_client
    SB_OK = True
except ImportError:
    SB_OK = False
    print("[ERROR] supabase-py not installed. Run: pip install supabase")

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False
    print("[WARN] pdfplumber not installed — CEA RE PDF parsing disabled. pip install pdfplumber")

try:
    import openpyxl
    XL_OK = True
except ImportError:
    XL_OK = False
    print("[WARN] openpyxl not installed — NPP XLS parsing disabled. pip install openpyxl")

# ── Config ────────────────────────────────────────────────────────────────
IST          = ZoneInfo("Asia/Kolkata")
SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_ANON_KEY')

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-7s  %(message)s",
                    datefmt="%H:%M:%S", handlers=[logging.StreamHandler(sys.stdout)])
log = logging.getLogger(__name__)

# State → region mapping
REGION_MAP = {
    'chandigarh':'Northern','delhi':'Northern','haryana':'Northern',
    'himachal pradesh':'Northern','jammu and kashmir':'Northern',
    'jammu & kashmir':'Northern','ladakh':'Northern','punjab':'Northern',
    'rajasthan':'Northern','uttar pradesh':'Northern','uttarakhand':'Northern',
    'chhattisgarh':'Western','gujarat':'Western','madhya pradesh':'Western',
    'maharashtra':'Western','daman and diu':'Western','goa':'Western',
    'dadra and nagar haveli':'Western',
    'andhra pradesh':'Southern','telangana':'Southern','karnataka':'Southern',
    'kerala':'Southern','tamil nadu':'Southern','pondicherry':'Southern',
    'puducherry':'Southern',
    'bihar':'Eastern','jharkhand':'Eastern','odisha':'Eastern',
    'west bengal':'Eastern','sikkim':'Eastern',
    'arunachal pradesh':'North-Eastern','assam':'North-Eastern',
    'manipur':'North-Eastern','meghalaya':'North-Eastern',
    'mizoram':'North-Eastern','nagaland':'North-Eastern','tripura':'North-Eastern',
}

KNOWN_STATES = set(REGION_MAP.keys())


# ─────────────────────────────────────────────────────────────────────────
# HELPER
# ─────────────────────────────────────────────────────────────────────────

def _find_col(df, candidates):
    cols = {c.lower().strip(): c for c in df.columns}
    for c in candidates:
        if c.lower() in cols:
            return cols[c.lower()]
        for cl, co in cols.items():
            if c.lower() in cl:
                return co
    return None


# ─────────────────────────────────────────────────────────────────────────
# PARSER 1 — CEA OPM Generation CSV
# URL: cea.nic.in/opm_grid_operation/daily-generation-report/
# Columns: Source, Value, Date & Time
# ─────────────────────────────────────────────────────────────────────────

def parse_generation_csv(path: Path, date: str, snapshot: str = 'eod') -> list:
    log.info(f"  Parsing CEA OPM generation CSV: {path.name}")
    df = pd.read_csv(path)
    df.columns = [c.strip().strip('"') for c in df.columns]

    src_col  = _find_col(df, ['source', 'Source', 'fuel', 'type'])
    val_col  = _find_col(df, ['value', 'Value', 'mw', 'generation'])
    time_col = _find_col(df, ['date & time', 'Date & Time', 'datetime', 'time'])

    if not all([src_col, val_col, time_col]):
        log.warning(f"  Could not identify columns. Found: {list(df.columns)}")
        log.warning("  Expected: Source, Value, Date & Time")
        return []

    df['_hour'] = pd.to_datetime(df[time_col].astype(str).str.strip(), errors='coerce').dt.hour
    df['_src']  = df[src_col].astype(str).str.strip().str.upper()
    df['_val']  = pd.to_numeric(df[val_col].astype(str).str.replace(',',''), errors='coerce')

    agg = df.groupby(['_src','_hour'])['_val'].mean().reset_index()
    records = []
    for _, r in agg.iterrows():
        if pd.isna(r['_val']): continue
        src = r['_src']
        if not src.endswith('GENERATION') and 'GENERATION' not in src:
            src = src + ' GENERATION'
        records.append({'data_date': date, 'hour': int(r['_hour']),
                        'source': src, 'value_mw': round(float(r['_val']),2),
                        'snapshot_time': snapshot})

    log.info(f"  ✓ {len(records)} generation rows parsed")
    return records


# ─────────────────────────────────────────────────────────────────────────
# PARSER 2 — CEA OPM Demand CSV
# URL: cea.nic.in/opm_grid_operation/daily-generation-report/
# Columns: Value, Time
# ─────────────────────────────────────────────────────────────────────────

def parse_demand_csv(path: Path, date: str, snapshot: str = 'eod') -> list:
    log.info(f"  Parsing CEA OPM demand CSV: {path.name}")
    df = pd.read_csv(path)
    df.columns = [c.strip().strip('"') for c in df.columns]

    val_col  = _find_col(df, ['value','Value','demand','mw','MW'])
    time_col = _find_col(df, ['time','Time','date & time','Date & Time','datetime'])

    if not all([val_col, time_col]):
        log.warning(f"  Could not identify columns. Found: {list(df.columns)}")
        return []

    df['_hour'] = pd.to_datetime(df[time_col].astype(str).str.strip(), errors='coerce').dt.hour
    df['_val']  = pd.to_numeric(df[val_col].astype(str).str.replace(',',''), errors='coerce')

    agg = df.groupby('_hour')['_val'].mean().reset_index()
    records = [{'data_date': date, 'hour': int(r['_hour']),
                'value_mw': round(float(r['_val']),2), 'snapshot_time': snapshot}
               for _, r in agg.iterrows() if not pd.isna(r['_val'])]

    log.info(f"  ✓ {len(records)} demand rows parsed")
    return records


# ─────────────────────────────────────────────────────────────────────────
# PARSER 3 — CEA RE Daily PDF (pdfplumber)
# URL: cea.nic.in/renewable-generation-report/
# Table: State | Wind MU | Solar MU | Others MU | Total MU
# ─────────────────────────────────────────────────────────────────────────

def parse_cea_re_pdf(path: Path, date: str) -> list:
    if not PDF_OK:
        log.error("  pdfplumber not installed. Run: pip install pdfplumber")
        return []

    log.info(f"  Parsing CEA RE PDF: {path.name}")
    records = []
    seen    = set()

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row or len(row) < 4: continue
                    cells   = [str(c or '').strip() for c in row]
                    raw     = cells[0].replace('\n',' ').strip().lower()

                    # Match state name
                    matched = None
                    for s in KNOWN_STATES:
                        if s in raw:
                            matched = s; break

                    # Check for region/all-india totals
                    is_ai = 'all india' in raw
                    is_rg = any(r in raw for r in
                        ['northern region','western region','southern region',
                         'eastern region','north-eastern region'])

                    if not matched and not is_ai and not is_rg: continue
                    key = matched or ('all_india' if is_ai else raw[:20])
                    if key in seen: continue

                    # Parse 4 numbers: Wind, Solar, Others, Total
                    nums = []
                    for cell in cells[1:]:
                        v = re.sub(r'[^\d.]','', cell.replace(',',''))
                        if v:
                            try: nums.append(float(v))
                            except ValueError: pass
                        if len(nums) == 4: break

                    if len(nums) < 3: continue
                    if all(n == 0 for n in nums[:3]): continue

                    seen.add(key)
                    total = nums[3] if len(nums) >= 4 else sum(nums[:3])
                    state = matched.title() if matched else ('All India' if is_ai else raw.title())
                    records.append({
                        'data_date': date,
                        'state':     state,
                        'region':    REGION_MAP.get(matched,'Other') if matched else ('All India' if is_ai else 'Region'),
                        'wind_mu':   nums[0], 'solar_mu': nums[1],
                        'others_mu': nums[2], 'total_mu': total,
                    })

    log.info(f"  ✓ {len(records)} state RE records parsed")
    if not records:
        log.warning("  No records extracted — PDF may use bilingual merged cells. Try a different page rotation or check pdfplumber table settings.")
    return records


# ─────────────────────────────────────────────────────────────────────────
# PARSER 4 — NPP dgr2 XLS (Station/Unit Wise)
# URL: npp.gov.in/publishedReports → dgr2-YYYY-MM-DD.xls (~382 KB)
# Most comprehensive daily file — Region > State > Sector > Type > Station > Unit
# ─────────────────────────────────────────────────────────────────────────

def parse_npp_dgr2(path: Path, date: str) -> tuple:
    if not XL_OK:
        log.error("  openpyxl not installed. Run: pip install openpyxl")
        return [], {}

    log.info(f"  Parsing NPP dgr2 XLS: {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)]
    wb.close()

    re_by_state  = {}
    fuel_totals  = {}
    cur_state = None
    cur_fuel  = None
    re_fuels  = {'wind','solar','small hydro','biomass','bagasse','others re','renewable'}

    for row in rows:
        if not row or row[0] is None: continue
        label   = str(row[0]).strip()
        label_l = label.lower()

        if label.title() in {s.title() for s in KNOWN_STATES}:
            cur_state = label.title()
            if cur_state not in re_by_state:
                re_by_state[cur_state] = {'wind':0,'solar':0,'others':0}

        for f in re_fuels:
            if f in label_l:
                cur_fuel = f; break
        else:
            if any(k in label_l for k in ['coal','thermal','nuclear','gas','diesel']):
                cur_fuel = None

        # Actual generation usually in col D (index 3)
        actual = None
        for ci in [3,4,2]:
            if ci < len(row) and row[ci] is not None:
                try:
                    actual = float(str(row[ci]).replace(',',''))
                    break
                except (ValueError, TypeError):
                    pass

        if actual and actual > 0:
            if cur_fuel:
                fuel_totals[cur_fuel] = fuel_totals.get(cur_fuel,0) + actual
            if cur_state and cur_fuel in re_fuels:
                if 'wind'  in cur_fuel: re_by_state[cur_state]['wind']   += actual
                elif 'solar' in cur_fuel: re_by_state[cur_state]['solar'] += actual
                else:                     re_by_state[cur_state]['others'] += actual

    state_records = []
    for state, v in re_by_state.items():
        if not any(v.values()): continue
        total = v['wind'] + v['solar'] + v['others']
        state_records.append({
            'data_date': date, 'state': state,
            'region':    REGION_MAP.get(state.lower(),'Other'),
            'wind_mu':   round(v['wind'],3),   'solar_mu':  round(v['solar'],3),
            'others_mu': round(v['others'],3), 'total_mu':  round(total,3),
        })

    log.info(f"  ✓ {len(state_records)} state records from dgr2. Fuel totals: {fuel_totals}")
    return state_records, fuel_totals


# ─────────────────────────────────────────────────────────────────────────
# PARSER 5 — NPP dgr3 XLS (All India Summary — 9 KB)
# ─────────────────────────────────────────────────────────────────────────

def parse_npp_dgr3(path: Path, date: str) -> dict:
    if not XL_OK: return {}
    log.info(f"  Parsing NPP dgr3 XLS: {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    fuel_map = {
        'coal':'thermal_mu','thermal':'thermal_mu','nuclear':'nuclear_mu',
        'hydro':'hydro_mu','gas':'gas_mu','wind':'wind_mu',
        'solar':'solar_mu','total':'total_mu',
    }
    figures = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None: continue
        label = str(row[0]).strip().lower()
        for fuel, key in fuel_map.items():
            if fuel in label:
                for ci in [3,4,2]:
                    try:
                        v = row[ci] if ci < len(row) else None
                        if v is not None:
                            figures[key] = round(float(str(v).replace(',','')),2)
                            break
                    except (ValueError, TypeError): pass
                break
    wb.close()
    log.info(f"  ✓ dgr3 figures: {figures}")
    return figures


# ─────────────────────────────────────────────────────────────────────────
# DAILY SUMMARY
# ─────────────────────────────────────────────────────────────────────────

def build_summary(date, gen_records, dem_records, sources, fuel_figs=None):
    def get(src):  return [r['value_mw'] for r in gen_records if src in r.get('source','')]
    def avg(lst):  return sum(lst)/len(lst) if lst else 0
    def mx(lst):   return max(lst) if lst else None

    solar  = get('SOLAR');  wind = get('WIND');  demand = [r['value_mw'] for r in dem_records]
    re_share = round((avg(solar)+avg(wind))/avg(demand)*100,1) if demand else None

    return {
        'data_date':         date,
        'peak_demand_mw':    mx(demand),
        'peak_solar_mw':     mx(solar),
        'peak_wind_mw':      mx(wind),
        'avg_re_share_pct':  re_share,
        'total_solar_mu':    round(avg(solar)*24/1000,3) if solar else (fuel_figs or {}).get('solar_mu'),
        'total_wind_mu':     round(avg(wind)*24/1000,3)  if wind  else (fuel_figs or {}).get('wind_mu'),
        'total_re_mu':       round((avg(solar)+avg(wind))*24/1000,3) if (solar or wind) else None,
        'total_demand_mu':   round(avg(demand)*24/1000,3) if demand else None,
        'data_sources':      sources,
        'notes':             f'Uploaded {datetime.now(IST).strftime("%Y-%m-%d %H:%M IST")} | ' +
                             (f'dgr3: {fuel_figs}' if fuel_figs else ''),
        'updated_at':        datetime.now(IST).isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────
# SUPABASE UPLOAD
# ─────────────────────────────────────────────────────────────────────────

def upload(date, gen=None, dem=None, re=None, summary=None):
    if not SB_OK: return
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)
    rows = 0

    if gen:
        sb.table('power_generation').upsert(gen, on_conflict='data_date,hour,source,snapshot_time').execute()
        rows += len(gen); log.info(f"  ✓ {len(gen)} generation rows → Supabase")
    if dem:
        sb.table('power_demand').upsert(dem, on_conflict='data_date,hour,snapshot_time').execute()
        rows += len(dem); log.info(f"  ✓ {len(dem)} demand rows → Supabase")
    if re:
        sb.table('power_re_state').upsert(re, on_conflict='data_date,state').execute()
        rows += len(re); log.info(f"  ✓ {len(re)} state RE rows → Supabase")
    if summary:
        sb.table('power_daily_summary').upsert(summary, on_conflict='data_date').execute()
        log.info(f"  ✓ Daily summary saved for {date}")
    sb.table('fetch_log').insert({'snapshot':'manual-upload','status':'ok',
        'sources': summary.get('data_sources',['manual']) if summary else ['manual'],
        'rows_written': rows}).execute()

    log.info(f"\n✓ Upload complete — {rows} total rows for {date}")


# ─────────────────────────────────────────────────────────────────────────
# AUTO-FIND FILES
# ─────────────────────────────────────────────────────────────────────────

def auto_find(date):
    dirs = [Path('.'), Path(f'power_data/{date}')]
    found = {}
    pats = {
        'gen':  [f'All_India_Generation*{date}*.csv', f'*eneration*{date}*.csv'],
        'dem':  [f'Demand_Met_Data*{date}*.csv',      f'*emand*{date}*.csv'],
        're':   [f'*re*{date}*.pdf',                   f'*RE*{date}*.pdf', f'*renewable*{date}*.pdf'],
        'dgr2': [f'dgr2*{date}*.xls',                  f'dgr2-{date}.xls'],
        'dgr3': [f'dgr3*{date}*.xls',                  f'dgr3-{date}.xls'],
    }
    for key, patterns in pats.items():
        for d in dirs:
            if not d.exists(): continue
            for p in patterns:
                m = list(d.glob(p))
                if m: found[key] = m[0]; log.info(f"  Auto-found {key}: {m[0]}"); break
            if key in found: break
    return found


# ─────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────

def main():
    default_date = (datetime.now(IST) - timedelta(days=1)).strftime('%Y-%m-%d')
    ap = argparse.ArgumentParser(description='Upload India power data to Supabase',
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--date',     default=default_date)
    ap.add_argument('--gen',      default=None, help='CEA OPM generation CSV')
    ap.add_argument('--dem',      default=None, help='CEA OPM demand CSV')
    ap.add_argument('--re',       default=None, help='CEA RE daily PDF')
    ap.add_argument('--dgr2',     default=None, help='NPP dgr2 XLS (station-level)')
    ap.add_argument('--dgr3',     default=None, help='NPP dgr3 XLS (summary)')
    ap.add_argument('--snapshot', default='eod', help='08:00 / 13:00 / 18:00 / eod')
    ap.add_argument('--auto',     action='store_true', help='Auto-find files for date')
    args = ap.parse_args()
    date = args.date

    log.info(f"\n{'='*55}")
    log.info(f"  India Power Analytics Uploader  |  {date}")
    log.info(f"  Sources: CEA OPM · CEA RE · NPP dgr2/dgr3")
    log.info(f"{'='*55}")

    if args.auto:
        log.info("\nAuto-scanning...")
        found = auto_find(date)
        for k in ['gen','dem','re','dgr2','dgr3']:
            if not getattr(args,k) and k in found:
                setattr(args, k, str(found[k]))

    gen_rows = []; dem_rows = []; re_rows = []; fuel_figs = {}; sources = []

    if args.gen:
        p = Path(args.gen)
        if p.exists(): gen_rows = parse_generation_csv(p, date, args.snapshot); sources.append('cea-opm-gen')
        else: log.error(f"Not found: {p}")

    if args.dem:
        p = Path(args.dem)
        if p.exists(): dem_rows = parse_demand_csv(p, date, args.snapshot); sources.append('cea-opm-dem')
        else: log.error(f"Not found: {p}")

    if args.re:
        p = Path(args.re)
        if p.exists(): re_rows = parse_cea_re_pdf(p, date); sources.append('cea-re-pdf')
        else: log.error(f"Not found: {p}")

    if args.dgr2:
        p = Path(args.dgr2)
        if p.exists():
            dgr2_re, dgr2_fuels = parse_npp_dgr2(p, date)
            existing = {r['state'] for r in re_rows}
            re_rows += [r for r in dgr2_re if r['state'] not in existing]
            fuel_figs.update(dgr2_fuels); sources.append('npp-dgr2')
        else: log.error(f"Not found: {p}")

    if args.dgr3:
        p = Path(args.dgr3)
        if p.exists():
            figs = parse_npp_dgr3(p, date)
            fuel_figs.update(figs)
            if figs: sources.append('npp-dgr3')
        else: log.error(f"Not found: {p}")

    if not any([gen_rows, dem_rows, re_rows, fuel_figs]):
        log.warning("\nNo data to upload. Specify --gen --dem --re --dgr2 --dgr3 or use --auto")
        ap.print_help(); return

    summary = build_summary(date, gen_rows, dem_rows, sources, fuel_figs) if sources else None
    log.info("\nUploading to Supabase...")
    upload(date, gen_rows or None, dem_rows or None, re_rows or None, summary)
    log.info(f"\n  Dashboard: https://india-power-analytics.vercel.app")


if __name__ == '__main__':
    main()
